# -*- coding: utf-8 -*-
"""


@author: Florian H


NOTE: how many regions/areas there are in the different countries
      For the calculation of the market values ​​it is assumed that there is only ONE region and area per country


-------------------
TODO:
- create plots and charts for typical variables (e.g. stacked area for generation, hydro storage, import/export)
- GUI ?
- extract 'list_of_REN' out of Input-Excel or other source? 



Based on previous work by Elaine T. Hale (https://pypi.org/project/gdxpds/) and
Philipp Andreas Gunkel (DTU).

    This program is free software: you can redistribute it and/or modify
    it under the terms of the GNU General Public License as published by
    the Free Software Foundation, either version 3 of the License, or
    (at your option) any later version.

    This program is distributed in the hope that it will be useful,
    but WITHOUT ANY WARRANTY; without even the implied warranty of
    MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
    GNU General Public License for more details.

    You should have received a copy of the GNU General Public License
    along with this program.  If not, see <https://www.gnu.org/licenses/>.

"""

#
# PREREQUISITES
#


# importing required packages and modules (keep order)
import os
from os import path
import config
import datetime
import glob
import gdxpds as gp
import pandas as pd
from openpyxl import load_workbook
import sys
from collections import defaultdict
import tkinter as tk
from tkinter import filedialog
import matplotlib.pyplot as plt
import inspect


# path to the local gams installation; change path in config.py
gams_dir_cache = config.local_gams

# project name; change in config.py
pr_name = config.project_name

# directory to the input gdx file(s)
gdx_file_list = glob.glob('input/*.gdx')

# create output directory
if not os.path.isdir('output'):
    os.mkdir('output')

# create directory for plots
if not os.path.isdir('plots'):
    os.mkdir('plots')
 
# time program start
time = datetime.datetime.now().strftime('%y%m%d-%H%M')

#default path MainResults
path_Main = os.getcwd() + '\\test_input\MainResults_flex.gdx'
#default path Baseresults
path_Base = os.getcwd() + '\\test_input\BASE-results_flex.gdx'

#market values are culculated for all of this technologies, if available
listofRENTech = ['K5_GNR_WT_WIND_ONSHORE_2020','K5_GNR_WT_WIND_OFFSHORE_2020','K5_GNR_WT_WIND_ONSHORE_2030',
                 'K5_GNR_WT_WIND_OFFSHORE_2030','K5_GNR_WT_WIND_ONSHORE_2040','K5_GNR_WT_WIND_OFFSHORE_2040',
                 'K5_GNR_WT_WIND_ONSHORE_2050','K5_GNR_WT_WIND_OFFSHORE_2050','K5_GNR_RES_WTR_PMP_2015',
                 'K5_GNR_RES_WTR_PMP_2020','K5_GNR_RES_WTR_PMP_2030','K5_GNR_RES_WTR_PMP_2040','K5_GNR_RES_WTR_PMP_2050',
                 'K5_GNR_ROR_WTR_2020','K5_GNR_ROR_WTR_2030','K5_GNR_ROR_WTR_2040','K5_GNR_ROR_WTR_2050',
                 'K5_GNR_RES_WTR_NOPMP_2020','K5_GNR_RES_WTR_NOPMP_2030','K5_GNR_RES_WTR_NOPMP_2040',
                 'K5_GNR_RES_WTR_NOPMP_2050','K5_GNR_PV_SUN_CENTRAL_2020','K5_GNR_PV_SUN_CENTRAL_2030',
                 'K5_GNR_PV_SUN_CENTRAL_2040','K5_GNR_PV_SUN_CENTRAL_2050','K5_GNR_PV_SUN_DECENTRAL_2020',
                 'K5_GNR_PV_SUN_DECENTRAL_2030','K5_GNR_PV_SUN_DECENTRAL_2040','K5_GNR_PV_SUN_DECENTRAL_2050',
                 'K5_GNR_ST_WOODCHIPS_EXT_2020','K5_GNR_ST_WOODCHIPS_EXT_2030','K5_GNR_ST_WOODCHIPS_EXT_2040',
                 'K5_GNR_ST_WOODCHIPS_EXT_2050','K5_GNR_ST_BIOGAS_EXT_2020','K5_GNR_ST_BIOGAS_EXT_2030',
                 'K5_GNR_ST_BIOGAS_EXT_2040','K5_GNR_ST_BIOGAS_EXT_2050','K5_GNR_GEOTHERMAL_2020',
                 'K5_GNR_GEOTHERMAL_2030','K5_GNR_GEOTHERMAL_2040','K5_GNR_GEOTHERMAL_2050','K5_GNR_HEAT_PUMP_EL',
                 'K5_GNR_BO_WOODCHIPS_2020','K5_GNR_BO_WOODCHIPS_2030','K5_GNR_BO_WOODCHIPS_2040',
                 'K5_GNR_BO_WOODCHIPS_2050','K5_GNR_BO_GEOTHERMAL_2020','K5_GNR_BO_GEOTHERMAL_2030',
                 'K5_GNR_BO_GEOTHERMAL_2040','K5_GNR_BO_GEOTHERMAL_2050','K5_GNR_HEAT_PUMP_EL_2020',
                 'K5_GNR_HEAT_PUMP_EL_2030','K5_GNR_HEAT_PUMP_EL_2040','K5_GNR_HEAT_PUMP_EL_2050',
                 'K5_GNR_BATTERY_LS_2020','K5_GNR_BATTERY_LS_2030','K5_GNR_BATTERY_LS_2040','K5_GNR_BATTERY_LS_2050']

def convertCountryRegion(in_conv,mode):
    # function converts Country into region and vice versa
    # input type: string, return value type: string
    # mode 0 = country into region
    # mode 1 = region into country
    
    CinSim = {'Country':['AUSTRIA','BELGIUM','BULGARIA','CROATIA','CYPRUS','CZECHIA','DENMARK','ESTONIA','FINLAND',
              'FRANCE','GERMANY','GREECE','HUNGARY','IRELAND','ITALY','LATVIA','LITHUANIA','LUXEMBOURG',
              'MALTA','NETHERLANDS','POLAND','PORTUGAL','ROMANIA','SLOVAKIA','SLOVENIA','SPAIN','SWEDEN',
              'SWITZERLAND','NORWAY','UNITED_KINGDOM']}
    RinSim = {'Region':['AT','BE','BG','HR','CY','CZ','DK','EE','FI','FR','DE','EL','HU','IE','IT','LV','LT','LU',
              'MT','NL','PO','PT','RO','SK','SI','ES','SE','CH','NO','UK']}
    
    #creating dataframes
    df1 = pd.DataFrame(data = CinSim)
    df2 = pd.DataFrame(data = RinSim)
    
    df3=df1.merge(df2, how='outer', left_index=True, right_index=True)
    
    if mode:
        out = df3.loc[df3['Region'] == in_conv]
        return out['Country'].values[0]
    else:
        out = df3.loc[df3['Country'] == in_conv]
        return out['Region'].values[0]
    

def get_data_price(df_price):
    #function gets dataframe with price information and returns dictonary with C as index
    #input: year specific dataframe with prices of all C in simulation
    #return value: dataframe dictonary; Country is index, columns are prices of regions per country
    
    all_price_c = df_price['C'].unique()        #getting all countries in dataframe
    df_price_collect_c = {}                     #declaring dictonary for countries
    df_price_collect_r = {}                     #declaring dictonary for regions
    
    for C in all_price_c:                       #loop over all countries in dataframe
        print('Country: ' + C)
        df_price_collect_c[C] = df_price.loc[df_price['C'].isin([C])]
        
        all_price_r = df_price_collect_c[C]['RRR'].unique()         #getting all regions per country
        
        for region in all_price_r:                  #loop over all regions
            print('Region: ' + region)
            df_price_collect_r[region]=df_price_collect_c[C].loc[df_price_collect_c[C]['RRR'].isin([region])]  #getting all data of specific region
            df_price_collect_r[region].rename(columns={'Value':region},inplace=True)                           #rename column name from Value to current region name
            df_price_collect_r[region].drop(['RRR'],axis=1,inplace=True)                                       #drop column 'RRR'
            

            if region in all_price_r[0]:        # for the first itteration, create temp dataframe and reset index

                temp = pd.DataFrame()           #create new dataframe
                temp = df_price_collect_r[region].drop(region,axis=1)   #drop column with values
                temp.reset_index(drop=True, inplace=True)               #reseting index
            
            df_price_collect_r[region].reset_index(drop=True, inplace=True)     #reset index to enable merging by index in the next line
            df_price_collect_c[C] = temp.merge(df_price_collect_r[region][region], how='outer', left_index=True, right_index=True) #merge temp and existing dataframe
            temp = df_price_collect_c[C]        
    
    return df_price_collect_c

def get_data_VGE(df_VGE):
    #function receives dataframe with generation data for one year and creates dataframes per area. columns are generation per technologie
    #input: dataframe with all VGE_T data which belongs to one year
    #return value: dictonary with area as index; columns of dataframes Year,Area,S,T,+all technologies
    
    #declaring dictonaries
    data_collect_area = {}
    data_collect_tech = {}
    data_collect_techfull = {}
    data_collect_merge = {}
    
    #get all areas
    all_areas_peryear = df_VGE['Area'].unique()
    for area in all_areas_peryear:                  #loop over all areas
        
        print('Area: ' + area)
        data_collect_area[area]=df_VGE.loc[df_VGE['Area'].isin([area])]         #assign area data to dictonary
        
        #get all Technologies per area
        all_tech_perarea = data_collect_area[area]['Generator'].unique()
        
        for technologie in all_tech_perarea:        #loop over all tech in area
            print('Technology: ' + technologie)
            data_collect_tech[technologie]=data_collect_area[area].loc[data_collect_area[area]['Generator'].isin([technologie])] #assign tech related data to tech dictonary

            if data_collect_tech[technologie].empty:    #if a dataframe is empty, don't add it to dictonary techfull
                print('Dataframe empty')
            else:
                data_collect_techfull[technologie]=data_collect_tech[technologie]
            
            data_collect_techfull[technologie].rename(columns={'Level':technologie},inplace=True) #rename column Level to technolgy name
            data_collect_techfull[technologie].drop('Generator',axis=1,inplace=True)              #drop column with technology name
            
            if technologie in all_tech_perarea[0]:          #if it's the first iteration, create temp dataframe
                #del temp
                temp = pd.DataFrame()                                                   #creating new dataframe
                temp = data_collect_techfull[technologie].drop(technologie,axis=1)      #drop tech data
                temp.reset_index(drop=True, inplace=True)                               #reseting index
                
            data_collect_techfull[technologie].reset_index(drop=True, inplace=True)     #reset index to enable merging by index in the next line
            data_collect_merge[area] = temp.merge(data_collect_techfull[technologie][technologie], how='outer', left_index=True, right_index=True) #merge temp and current dataframe
            temp = data_collect_merge[area]
          
    return data_collect_merge                              #return value: dictonary with area as index; columns of dataframes Year,Area,S,T,+all technologies

    
def get_data_QEEQ(df_QEEQ):
    #function receives dataframe with QEEQ data for one year and creates dataframes with all data from all regions in columns
    #input: dataframe with all QEEQ data which belongs to one year
    #return value: dataframe with all QEEQ data. regions as columns; columns: Year,S,T,+all regions      
    
    #declaring dictonaries
    data_collect_region = {}
    data_collect_merge = {}
    
    
    all_region_peryear = df_QEEQ['Region'].unique()  #getting all regions in QEEQ
    
    for region in all_region_peryear:       #loop over all regions
        
        print('Region: ' + region)
        data_collect_region[region]=df_QEEQ.loc[df_QEEQ['Region'].isin([region])]           #split data into regions and add to dictonary
        data_collect_region[region].rename(columns={'Level':region},inplace=True)           #rename column Level to name of area
        data_collect_region[region].drop('Region',axis=1,inplace=True)                      #drop column region
        
        if region in all_region_peryear[0]:     #if it's the first iteration, create temp dataframe
            #del temp
            temp = pd.DataFrame()
            temp = data_collect_region[region].drop(region,axis=1)
            temp.reset_index(drop=True, inplace=True)
        
        data_collect_region[region].reset_index(drop=True, inplace=True)     #reset index to enable merging by index in the next line
        data_collect_merge= temp.merge(data_collect_region[region][region], how='outer', left_index=True, right_index=True)     #merge temp and current dataframe
        temp = data_collect_merge
        
    print('End of QEEQ')
    return data_collect_merge

def isRENTech(Tech):
    #check if tech is a renewable; see listofRENTech; if technologie is found in listofRENTech return is TRUE
    #input: name of technologie
    #return: boolean (True/False)
    
    if Tech in listofRENTech:
        return True
    else:
        return False
   

def culc_Total_Gen_tech(df_VGE_tech):
    #culcalate tech specific total generation
    #input: dictonary with VGE_T data for one year; area is index of dictonary
    #return value: dictonary with area as index; columns of dataframes: Technologies; value is total generation of technologie 
    
    df_Gen_tech_sum = pd.DataFrame()                #create new dataframe
    
    df_collect = {}                                 #declaring dictonary
    list_of_areas = list(df_VGE_tech.keys())        #getting list of all indices 
    
    for area in list_of_areas:                      #loop over all areas
        list_df_header = list(df_VGE_tech[area].columns)            #get all technologies
        
        del df_Gen_tech_sum                         #delete dataframe bevor creating a new one
        df_Gen_tech_sum = pd.DataFrame()            #create new dataframe
        
        for tech in list_df_header[4:]:             #loop over technologies; technologies start at 4th column         
            if isRENTech(tech):                     #checking for renewable technologie; only market values from renewable techs are culculated
                sum_tech=df_VGE_tech[area][tech].sum(axis=0)        #culcalating sum 
                df_Gen_tech_sum[tech] = None
                df_Gen_tech_sum.at[0,tech]= sum_tech                #add entry in line 0, column tech
        df_collect[area] = df_Gen_tech_sum                          #add dataframe to dictonary
        
    return df_collect

def culc_Rev_tech(df_price,df_VGE_tech):
    #culcalate tech specific revenue
    #input: 
    #return: 
    
    list_of_countries = list(df_price.keys())
    list_of_areas = list(df_VGE_tech.keys())
    
    tech_rev = pd.DataFrame()
    
    df_collect = {}
    df_collect_c = nested_dict(2, float)

    #culculate total revenue
    for country in list_of_countries:
        
        #convert country to region, mode 0
        region = convertCountryRegion(country,0)
        
        for area in list_of_areas:
            if area[:2]== region:
                list_df_header = list(df_VGE_tech[area].columns)
                for tech in list_df_header[4:]:
                    if isRENTech(tech):
                        rev_tech = df_price[country][region]*df_VGE_tech[area][tech]
                        rev_tech_sum = rev_tech.sum(axis=0)
                        tech_rev[tech] = None
                        tech_rev.at[0,tech]= rev_tech_sum
                        
                df_collect[area] = tech_rev
                del tech_rev
                tech_rev = pd.DataFrame()
        
                df_collect_c[country][area]=df_collect[area]
            #else:
                #print('BREAK')
    
    return df_collect_c

def nested_dict(n, type):
    #creates multidimensional dictonarys
    if n == 1:
        return defaultdict(type)
    else:
        return defaultdict(lambda: nested_dict(n-1, type))
    
def culc_market_val(price_avr,tech_gen,tech_rev):
    # culculate relativ market value for each tech
    print('culcalating market value')
    #get all regions (=countries)
    list_df_header = list(price_avr.columns)
    
    market_val = pd.DataFrame()
    
    data_collect_a = {}
    #create 2-dim dictonary
    data_collect_r = nested_dict(2, float)
    
    for region in list_df_header:
        country = convertCountryRegion(region,1)
        
        #get all areas in region
        list_of_areas = list(tech_rev[country].keys())

        for area in list_of_areas:
            #get tech in area
            list_of_tech = list(tech_rev[country][area].columns)
            for tech in list_of_tech:
                mkt_val = tech_rev[country][area][tech]/(tech_gen[area][tech])/price_avr[region]
                #building dataframe
                market_val[tech] = None
                market_val.at[0,tech]= mkt_val.values[0]
                
            data_collect_a[area] = market_val
            del market_val
            market_val = pd.DataFrame()
            
            data_collect_r[country][area]= data_collect_a[area]

    return data_collect_r

def print_results_df(year,sheetname,data):
    
    print('Writing Results to Excel ',year)
    csv_file = 'output/Results_'+ year +'_'+ time +'.xlsx'
    #creating excel file if it not exists
    if not path.exists(csv_file):
        null = pd.DataFrame()
        #create results excel-file           
        with pd.ExcelWriter(csv_file, engine="openpyxl") as writer:
            null.to_excel(writer, 'temp', encoding='utf8', index=False)
    
    #df2 = pd.DataFrame({'Data': [13, 24, 35, 46]})
    
    writer = pd.ExcelWriter(csv_file, engine='openpyxl',mode="a")

    data.to_excel(writer, sheetname, startcol=1,startrow=1)
    writer.save()
    # wb = load_workbook(csv_file)
    # if 'temp' in wb.sheetnames:
    #     del wb['temp']
    # wb.save(csv_file)

def print_results_dict(year,sheetname,data):
    
    print('Writing Results to Excel ',year)
    csv_file = 'output/Results_'+ year +'_'+ time +'.xlsx'
    #creating excel file if it not exists
    if not path.exists(csv_file):
        null = pd.DataFrame()
        #create results excel-file           
        with pd.ExcelWriter(csv_file, engine="openpyxl") as writer:
            null.to_excel(writer, 'temp', encoding='utf8', index=False)
    
    #df2 = pd.DataFrame({'Data': [13, 24, 35, 46]})
    
    writer = pd.ExcelWriter(csv_file, engine='openpyxl',mode="a")
    
    list_of_keys = data.keys()
    for key in list_of_keys:
        data[key].to_excel(writer, sheetname + key, startcol=1,startrow=1)
    
    
    writer.save()
    # wb = load_workbook(csv_file)
    # if 'temp' in wb.sheetnames:
    #     del wb['temp']
    # wb.save(csv_file)
    
    
def Program1(print_all):
    # function culcalates total generation, generation per technologie and market values
    # if print_all == 1 all variables are exported to excel-files. NOTE: This could take a long time
    # if print_all == 0 only a summary is exportet to excel-file
    # input value: boolean
    # return value: none
    
    df_return = {}          #declaring dictonary
    
    df_raw_price = proceedMainResults('1') #get price data

    print('----------------------------------------------------------------------------------------------------\n')
    
    df_return = proceedBaseResults('1')     #get generation data VGE_T and data QEEQ
    
    #assigning dataframes to seperate variables; 
    df_raw_VGE = df_return[0]                  
    df_raw_QEEQ = df_return[1]
    
    #check whether years in variables match
    allyears_price = df_raw_price[0]['Y'].unique()
    allyears_VGE = df_raw_VGE['Year'].unique()
    allyears_QEEQ = df_raw_QEEQ['Year'].unique()
        
    if not (len(allyears_price) == len(allyears_VGE) and len(allyears_VGE) == len(allyears_QEEQ) and len(allyears_price) == len(allyears_QEEQ)):
        print('Error: years dont match')
        sys.exit("Error message")
    else:        
        for element in range(0,len(allyears_price)):
            if not allyears_price[element] == allyears_VGE[element]:
                print('Error: years dont match')
                sys.exit("Error message")
    
    #declaration of dictonaries
    data_collect_price = {}
    data_collect_VGE = {}
    data_collect_QEEQ = {}
    
    data_collect_TotGen = {}
    data_collect_TotRev = {}
    data_collect_Price_average = {}
    
    data_collect_Gen = {}
    data_collect_Rev = {}
    data_collect_market_val = {}
    
    data_collect_merge = {}
    data_collect_area_merge = {}
    
    #getting data
    for year in allyears_price:     #loop over all years
        
        #getting price data
        df=df_raw_price[0].loc[df_raw_price[0]['Y'].isin([year])]       #assign all data from current year to new dataframe
        data_collect_price[year]=get_data_price(df)                     #formatting dataframe
        
        #getting generation data
        data_collect_VGE[year] = get_data_VGE(df_raw_VGE.loc[df_raw_VGE['Year'].isin([year])])  #getting all data from current year and format dataframe
        
        #getting total generation data
        data_collect_QEEQ[year] = get_data_QEEQ(df_raw_QEEQ.loc[df_raw_QEEQ['Year'].isin([year])])    #getting all data from current year and format dataframe
    
    # culculation total generation
    
    for year in allyears_price:
        data_collect_TotGen[year] = culc_Total_Gen(data_collect_VGE[year])
        print('TotGen: ',data_collect_TotGen[year])
        data_collect_TotRev[year] = culc_Total_Rev(data_collect_price[year],data_collect_TotGen[year])
        print('TotRev: ',data_collect_TotRev[year])
        data_collect_Price_average[year] = data_collect_TotRev[year]/data_collect_TotGen[year]
        print('Price_average: ',data_collect_Price_average[year])
        
        
    #culculation tech
    
    for year in allyears_price:   #loop over all years
        data_collect_Gen[year] = culc_Total_Gen_tech(data_collect_VGE[year])       #get tech specific generation
        print('TotGen_tech: ',data_collect_Gen[year])

        data_collect_Rev[year] = culc_Rev_tech(data_collect_price[year],data_collect_VGE[year])     #get tech specific revenue
        print('TotRev_tech: ',data_collect_Rev[year])
        input('STOP')
        #tech specific market value        
        data_collect_market_val[year] = culc_market_val(data_collect_Price_average[year],data_collect_Gen[year],data_collect_Rev[year])
        print('Market values: ',data_collect_market_val[year])
    
    # for year in allyears_price:
    #     print_results(year,data_collect_TotGen[year],data_collect_TotRev[year],data_collect_Price_average[year],data_collect_Gen[year],data_collect_Rev[year],data_collect_market_val[year])
    
    
    #merge results
    for year in allyears_price:
        data_collect_merge[year] = pd.concat([data_collect_TotGen[year],data_collect_TotRev[year],data_collect_Price_average[year]],ignore_index=True)
        #print(data_collect_merge[year])
        data_collect_merge[year].insert(0,'Type',['Total Gen','Total Rev','Avarage Price'])
        data_collect_merge[year].insert(0,'Tech',['All','All','All'])
        print('Result merge')
        
        
        #assumption only 1 region and 1 area per country
        list_of_region = list(data_collect_Price_average[year].columns)

        #create dataframe for tech gen
        df3 = pd.DataFrame(columns=['Tech','Type']+list_of_region)
        
        
        for region in list_of_region:
            country = convertCountryRegion(region,1)
        
            #get all areas in region
            list_of_areas = list(data_collect_Rev[year][country].keys())
            #print('AREAS')
            #print(list_of_areas)
            
            list_of_tech = list(data_collect_Gen[year][list_of_areas[0]].columns)
            #print('Tech ',list_of_tech, 'Year ', year)
            
            for tech in list_of_tech:
                #merge tech specific gen
                df = pd.DataFrame(columns=['Tech','Type',region])
                df.at[0, 'Tech']=tech
                df.at[0,'Type']='Generation'
                df[region]=data_collect_Gen[year][list_of_areas[0]][tech]
                if tech == list_of_tech[0]:
                    #print('FIRST')
                    data_collect_area_merge[year] = df
                    #print(data_collect_area_merge)
                else:
                    #print('ELSE')
                    data_collect_area_merge[year] = pd.concat([data_collect_area_merge[year],df],ignore_index=True)
                
                
                #merge tech specific revenue
                df2 = pd.DataFrame(columns=['Tech','Type',region])
                df2.at[0, 'Tech']=tech
                df2.at[0,'Type']='Revenue'
                df2[region]=data_collect_Rev[year][country][list_of_areas[0]][tech]
                
                data_collect_area_merge[year] = pd.concat([ data_collect_area_merge[year],df2],ignore_index=True)
                
                #merge tech specific market value
                df3 = pd.DataFrame(columns=['Tech','Type',region])
                df3.at[0, 'Tech']=tech
                df3.at[0,'Type']='Market Value'
                df3[region]=data_collect_market_val[year][country][list_of_areas[0]][tech]
                
                data_collect_area_merge[year] = pd.concat([ data_collect_area_merge[year],df3],ignore_index=True)
            
            if region == list_of_region[0]:
                df4 = pd.DataFrame(columns = ['Tech','Type'])
            df4=pd.merge(df4,data_collect_area_merge[year],on=['Tech','Type'],how='outer')

            
        data_collect_merge[year] = pd.concat([data_collect_merge[year],df4],ignore_index=True)
        del df4
            

    print('Total generation')
    print(data_collect_TotGen)
    print('Total revenue')
    print(data_collect_TotRev)
    print('Price average')
    print(data_collect_Price_average)
    for year in allyears_price:
        print_results_df(year, 'Results', data_collect_merge[year])
    
    if print_all:
        for year in allyears_price:
            print('PRINT QEEQ')
            print_results_df(year, 'QEEQ', data_collect_QEEQ[year])
            
            print('PRINT PRICE ',year)            
            print_results_dict(year, 'PRICE ',data_collect_price[year])
            
            print('PRINT VGE_T ',year)
            print_results_dict(year, 'EL_GEN ', data_collect_VGE[year])
    

def proceedMainResults(opt_main):
#Opening MainResults.gdx; Choose variable; get data; write it to excel or csv
    #opening MainResults.gdx in current working directory
    #opt_main: option which variable should be loaded from MainResults.gdx
    #opt_main == 3: chose variable
    #opt_main != 3: loading EL_PRICE_YCRST
    #return value: dictonary with number corresponding to variable as index; columns of dataframe Y,C,+not dropped columns
    
    print('Loading MainResults.gdx')
    gdx_file = path_Main                #new name for path
    #get wanted variable
    dataframes ={}                      #declaring dictonary
    
    if opt_main == '3':             #manually chose variable from MainResults.gdx
        variable_input = input('Variable?\n')
        variable_input_list = variable_input.split()        #input splitted by space
        #check for wrong input
        for i in range(0,len(variable_input_list)):         #looping over all variables
            dataframes[variable_input_list[i]] = gp.to_dataframe(gdx_file,variable_input_list[i], gams_dir = gams_dir_cache,old_interface=False)

        df_raw_collect = {}
        
        #loop through variable input
        for j in range(0,len(variable_input_list)):
            #get selcted dataframe
            df_raw_collect[j] = dataframes[variable_input_list[j]]
            #choose parameter
            list_df_header = list(df_raw_collect[j].columns)
            print(list_df_header[2:])                     #First two columns are always Y and C
            var_drop = input('Select from the above list which parameter you want to drop from variable ' + variable_input_list[j] + '\n (Year Y and country C are always mentioned)\n')
            var_drop_list = var_drop.split()    #input of columns to drop; splitted by space
        
            #checking wrong input
            for var in var_drop_list:
                if not var in list_df_header[2:]:
                    print('Error: variable not found in MainResults.gdx')
                    input('Press Enter to exit')
                    sys.exit("Error message")
            
            #dropping columns
            for k in range(0,len(var_drop_list)):
                df_raw_collect[j].drop(var_drop_list[k],axis=1,inplace=True)
            
            #replace EPS (=5e+300) with 0
            df_raw_collect[j]['Value'].values[df_raw_collect[j]['Value'].values > 4e300] = 0
        
        del dataframes            #deleting dataframes to free system memory
        
    else:                       #get data from variable EL_PRICE_YCRST; similar to opt_main == '3'    
        variable_input_list=['EL_PRICE_YCRST']
        var_drop_list=['UNITS']

        dataframes = gp.to_dataframe(gdx_file,variable_input_list[0], gams_dir = gams_dir_cache,old_interface=False)
        
        df_raw_collect = {}     #declaring dictonary
        #loop through variable input
        for j in range(0,len(variable_input_list)):
            #get selcted dataframe
            df_raw_collect[j] = dataframes
            #choose parameter
            list_df_header = list(df_raw_collect[j].columns)
        
            #checking wrong input
            for var in var_drop_list:
                if not var in list_df_header[2:]:
                    print('Error: variable not found in MainResults.gdx')
                    input('Press Enter to exit')
                    sys.exit("Error message")
                    
            #dropping columns
            for k in range(0,len(var_drop_list)):
                df_raw_collect[j].drop(var_drop_list[k],axis=1,inplace=True)
            
            #replace EPS (=5e+300) with 0
            df_raw_collect[j]['Value'].values[df_raw_collect[j]['Value'].values > 4e300] = 0
        
        del dataframes
    
    return(df_raw_collect)      #return dictonary with number corresponding to variable as index


def VGE(data_raw):
    #formating VGE_T; renaming and reordering of columns
    #input: raw VGE_T dataframe
    #return value: formated dataframe; columns Year,Area,Generation,S,T,Level
    
    #rename unnamed columns
    col_rename_list = ['Sce','Y1','Year','Area','Generator','S','T']
    for m in range(0,len(data_raw['*'].columns)):
        data_raw.columns.values[m] = col_rename_list[m]

    # drop unnecessary columns
    data_raw.drop(['Sce','Y1','Marginal','Lower','Upper','Scale'],axis=1,inplace=True)
    
    #replace EPS (=5e+300) with 0
    data_raw['Level'].values[data_raw['Level'].values > 4e300] = 0
    return data_raw


def QEEQ(df_QE_raw):
    #rename unnamed columns
    col_rename_list = ['Sce','Y1','Year','Region','S','T']
    for m in range(0,len(df_QE_raw['*'].columns)):
        df_QE_raw.columns.values[m] = col_rename_list[m]

    # drop unnecessary columns
    df_QE_raw.drop(['Sce','Y1','Marginal','Lower','Upper','Scale'],axis=1,inplace=True)
    
    #replace EPS (=5e+300) with 0
    df_QE_raw['Level'].values[df_QE_raw['Level'].values > 4e300] = 0
    
    return df_QE_raw
    
        
    
def proceedBaseResults(opt_res):
    #opening BaseResults, choosing variable, return dataframe
    #option '1': get VGE_T and QEEQ
    #option '2': get VGE_T
    #option '4': Manually choose variable
    #return value: dictonary of dataframes; dictonary with number corresponding to variable as index; dataframe columns
    #              from option or manually chosen
    
    
    df_return = {}      #declaring dictonary
    
    if opt_res == '1':          #option1: getting VGE_T and QEEQ
        print('Loading BaseResults.gdx')
        gdx_file = path_Base
        dataframes = gp.to_dataframe(gdx_file,'VGE_T', gams_dir = gams_dir_cache,old_interface=False)       #getting variable from .gdx to dataframe            
        df_return[0]=VGE(dataframes)                                                                        #processing data to wanted dataformat, see function VGE
        dataframes = gp.to_dataframe(gdx_file,'QEEQ', gams_dir = gams_dir_cache,old_interface=False)
        df_return[1]=QEEQ(dataframes)                                                                       #processing data to wanted dataformat, see function QEEQ
        
        del dataframes           #deleting dataframe to free memory space
        return df_return
    
    elif opt_res == '2':       #option1: getting VGE_T
        print('Loading BaseResults.gdx')
        gdx_file = path_Base
        dataframes = gp.to_dataframe(gdx_file,'VGE_T', gams_dir = gams_dir_cache,old_interface=False)
        df_return=VGE(dataframes)
        
        del dataframes
        return df_return
        
    elif opt_res == '4':
        print('Loading BaseResults.gdx')
        gdx_file = path_Base
        dataframes ={}
        #input wanted variables
        variable_input = input('Variable?\n')
        variable_input_list = variable_input.split()
        #get variables from *gdx
                   
        df_BR_raw_collect = {}

        
        #check for wrong input
        for l in range(0,len(variable_input_list)):
            dataframes[variable_input_list[l]] = gp.to_dataframe(gdx_file,variable_input_list[l], gams_dir = gams_dir_cache,old_interface=False)
                
                
            df_BR_raw_collect[l] = dataframes[variable_input_list[l]]
                
            #rename unnamed rows
            if len(df_BR_raw_collect[l] ['*']) > 0:

                #print(len(df_BR_raw_collect[l] ['*'].columns))
                print(len(df_BR_raw_collect[l] ['*'].columns),' columns of variable ' + variable_input_list[l] + ' are unnamned\n')
                col_rename = input('Please enter the names of the columns to rename them\n')
                col_rename_list = col_rename.split()
                
                print(col_rename_list)
                
                star_len = len(df_BR_raw_collect[l]['*'].columns)

                df_BR_raw_collect[l].columns=col_rename_list + list(df_BR_raw_collect[l].columns[star_len:])
                print(df_BR_raw_collect[l])
                
                list_df_header = list(df_BR_raw_collect[l].columns)
                print(list_df_header)
                var_drop = input('Select from the above list which parameter you want to drop from variable ' + variable_input_list[l] +'\n')
                var_drop_list = var_drop.split()
            
                #checking wrong input
                for var in var_drop_list:
                    if not var in list_df_header:
                        print('Error: variable not found in Base-results.gdx')
                        input('Press Enter to exit')
                        sys.exit("Error message")
                        
                    #df_BR_raw_collect[l].drop(df_BR_raw_collect[l].columns[3],axis=1,inplace=True)
                    df_BR_raw_collect[l].drop(var,axis=1,inplace=True)
                    print(df_BR_raw_collect[l])
                
            
            
            df_return[l] = df_BR_raw_collect[l]
        
        del dataframes
        return df_return
    else:
        input('Error: Wrong option BaseResults')
        sys.exit("Stopped (proceedBaseResults)")
    
        
    

def culc_Total_Gen(df_vge):
    # culculats total gen in region, return dataframe over all regions
    #input: dictonary with VGE_T data for one year; area is index of dictonary
    #return value:
    
    #declaring dictonaries
    gen_tot = pd.DataFrame()
    gen_temp = pd.DataFrame()
    gen = pd.DataFrame()
    
    list_of_areas = list(df_vge.keys())         #getting all areas from indices of dictonary
    
    #summerize total generation of all tech in area
    for area in list_of_areas:                  #loop over all areas
        sum_temp = df_vge[area].sum(axis=1)
    
        gen_temp[area] = sum_temp
    
    
    all_columns = gen_temp.columns

    #summerize over all areas of one country
    for areas in all_columns:               #loop over all areas
        
        if areas[2:] == '_A':
            print(areas[2:])
            input('STOP')
            #get all areas in the region
            df2 = gen_temp.filter(regex=areas[:2])
            #sum over all areas und name the columns as regions
            gen[areas[:2]] = df2.sum(axis=1)
            gen_temp2 = gen.sum(axis=0)
            new_region = areas[:2]
            gen_tot[new_region] = None
            gen_tot.at[0,new_region] = gen_temp2[new_region]

    return gen_tot


def culc_Total_Rev(df_price,df_tot_gen):
    # culcats total revenue, return dataframe
    
    print('Start culc_Total_Rev')
    print(df_price)
    list_of_countries = list(df_price.keys())

    tot_rev = pd.DataFrame()

    #culculate total revenue
    for country in list_of_countries:
        
        #convert country to region, mode 0
        region = convertCountryRegion(country,0)
        
        temp_price = df_price[country]

        temp_rev = temp_price[region]*df_tot_gen[region]  
        temp_rev_sum = temp_rev.sum(axis=0)
        new_country = region

        tot_rev[new_country] = None
        tot_rev.at[0,new_country] = temp_rev_sum
        
    
    print(tot_rev)
    #input('END')
    return tot_rev


def culculation(dict_main,dict_qeeq):            
    #culculation of results
    
    print(dict_main)
    print(dict_qeeq)
    
    d =[('Total Generation [MWh]')]
    gen_tot_c = pd.DataFrame(data = d, columns=['Type'])
    print(gen_tot_c)
    
    df_return = culc_Total_Gen(gen_tot_c,dict_qeeq)
    print(df_return)
    
    input('Total Gen end. Press Enter to continue')
    
    df_return = culc_Total_Rev(gen_tot_c,dict_main,dict_qeeq)
    print(df_return)
    input('Press Enter to continue')

    temp_rev_collection = {}
    #culculate total revenue
    for country in list_of_countries:
        temp_price = dict_main[country]
        
        for years in list_of_years:
            temp_gen_T = dict_qeeq[years]
            print(temp_price['AT'])
            print(temp_gen_T['AT'])
            temp_rev_collection[years] = temp_price.iloc[:,4]*temp_gen_T.iloc[:,3]
            print(temp_rev_collection[years])
            temp_sum = temp_rev_collection[years].sum(axis=0)
            new_row = {'Type':'Revenue','AT':temp_sum}
            gen_tot_c = gen_tot_c.append(new_row, ignore_index=True)

    print(gen_tot_c)
    # for years in list_of_keys:
    #    gen_tot
    input('Press Enter to continue')
    
    print('Starting printing of results to excel')
    
    
    df_results = df_collect_merge[country].sum(numeric_only=True)
    print(df_results)
    
    book = load_workbook(csv_file)
    writer = pd.ExcelWriter(csv_file, engine='openpyxl') 
    writer.book = book
    
    ## ExcelWriter for some reason uses writer.sheets to access the sheet.
    ## If you leave it empty it will not know that sheet Main is already there
    ## and will create a new sheet.
    
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
    
    df_results.to_excel(writer, "Results")
    
    writer.save()

def plot_VGE(aggregate):
    # plots VGE_T in a stacked area chart
    
    list_of_fuels = ['NGAS','COAL','LIGN','NUCL','WIND','SUN','WTR','BIOGAS','MSW','GEOTHERMAL','WOODCHIPS']
    df_return = proceedBaseResults('2')
    allyears = df_return['Year'].unique() 
    print(allyears)
    
    data_collect_VGE = {}
    
    for year in allyears:    
        data_collect_VGE[year] = get_data_VGE(df_return.loc[df_return['Year'].isin([year])])
        
        allareas = data_collect_VGE[year].keys()
        
        for area in allareas:
            if aggregate:
                df_temp = pd.DataFrame()
                for fuel in list_of_fuels:
                    
                    df = data_collect_VGE[year][area].filter(regex=fuel)   #filter columns for all tech with NGAS
                    sum_fuel = df.sum(axis=1)
                    df_fuel = sum_fuel.to_frame(name=fuel)
                    

                    
                    df_merge=df_temp.merge(df_fuel, how='outer', left_index=True, right_index=True)
                    df_temp = df_merge
                    
                    
                
                #df_merge.drop(columns=['COAL','LIGN','NUCL','MSW','GEOTHERMAL','WOODCHIPS','BIOGAS'],inplace=True)        
                print(df_merge)
               
                fig, ax = plt.subplots()
                ax = plt.stackplot(df_merge.index,df_merge.values.T)
                plt.xlim(4368,4536)
                plt.legend(df_merge.columns,loc='upper left')
                #plt.show()
                
                print('saving plot as pdf')
                fig.savefig('plots/'+year +'_'+ area +'_'+ 'VGE_T'+'_'+time+'.pdf')
            
            else:
                print('Start else \n')
                data_collect_VGE[year][area].drop(columns=['Year','Area','S','T'],inplace=True)
                print(data_collect_VGE[year][area])
                fig, ax = plt.subplots()
                ax = plt.stackplot(data_collect_VGE[year][area].index,data_collect_VGE[year][area].values.T)
                plt.xlim(4368,4536)
                plt.legend(data_collect_VGE[year][area].columns,loc='upper left')
                #plt.show()
                
                print('saving plot as pdf')
                fig.savefig('plots/'+year +'_'+ area +'_'+ 'VGE_T'+'_'+time+'.pdf')
        

    
    
#======================MAIN===================================================
#User interface

#look for input files
root = tk.Tk()
root.withdraw()

#print selected paths for .gdx-files
print('----------------------------------------------------------------------------------------------------\n')
print('Current paths of file MainResults')
print(path_Main)
print('----------------------------------------------------------------------------------------------------\n')
print('Current paths of file BaseResults')
print(path_Base)

#choose option
print('----------------------------------------------------------------------------------------------------\n')
print('Output Handling MainResults.gdx and BASE_results.gdx to .xlsx')
option = 99                 #arbitrary number to start the while loop in any case
while option != 0:          #end loop if option 0 is chosen
    
    print('\n Choose Option: (by number or character)\n')
    print('c: Choose new filepaths of MainResults and BaseResults\n')
    print('1:  calculate market values and export .xlsx')
    print('2:  calculate market values (Input for GREENX) and EL_Price + VGE_T + QEEQ to excel') 
    print('3:  Manually choose variable from MainResults and export to .xlsx')
    print('4:  Manually choose variable from Base-results and export to .xlsx')
    print('5:  Get electricity price for all countries and export to .xlsx')
    print('6:  Plot VGE_T in stacked area plot')
    option = input('0: Stop execution \n')
    print('----------------------------------------------------------------------------------------------------\n')
    
    if option == '1':       #start program 1 with option FALSE
        Program1(False)         
        
    elif option == '2':
        Program1(True)
    elif option == '3':
        df_return = proceedMainResults(option)
        for i in df_return.keys():
            print_results_df('0', str(i+1), df_return[i])
    elif option == '4':
        df_return = proceedBaseResults(option)
        print('DF RETURN')
        print(df_return)
        for i in df_return.keys():
            print_results_df('0', str(i+1), df_return[i])
    elif option == '5':       
        data_collect_price ={}
        df_price_raw = proceedMainResults('1')
        
        allyears_price = df_price_raw[0]['Y'].unique()
        for year in allyears_price:        
            #getting price data
            df=df_price_raw[0].loc[df_price_raw[0]['Y'].isin([year])]
            data_collect_price[year]=get_data_price(df)
            
        for year in allyears_price:
            print_results_dict(year, '', data_collect_price[year])
    
    elif option == '6':
        plot_VGE(False)
             
    elif option == 'c':
        print('Select MainResults-File')
        path_Main = filedialog.askopenfilename(title='Selcect MainResults.gdx',filetypes=[('gdx files','*.gdx'),('All files','*.*')])
        print(path_Main)
        
        print('Select BaseResults-File')
        path_Base = filedialog.askopenfilename(title='Selcect BASEResults.gdx',filetypes=[('gdx files','*.gdx'),('All files','*.*')])
        print(path_Base)
        print('----------------------------------------------------------------------------------------------------\n')
            
    else:
        sys.exit("End of execution.")
    
#proceedMainResults()
#
#input('MainResults end. Press Enter to continue')
#proceedBaseResults()
#input('BaseResults end. Press Enter to continue')
#print final statement
print('End of execution.')

if __name__ == '__main__':
    main()

# = END =======================================================================
